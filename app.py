#!/usr/bin/env python3
"""
PDF Data Extractor — Multi-User Web App
==========================================
Upload financial PDFs → auto-split large files → OCR → Extract to Excel.

PDFs with more than 10 pages are automatically split into 10-page chunks
before OCR processing to prevent memory issues.

DEPLOY:
  Streamlit Cloud: Push to GitHub, connect at share.streamlit.io
  Local:           streamlit run app.py
"""

import io
import os
import re
import tempfile
import time
import zipfile
from datetime import datetime
from math import ceil

import streamlit as st
from pdf2image import convert_from_path
import pytesseract
from PyPDF2 import PdfReader, PdfWriter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd


# =============================================================================
# PAGE CONFIG
# =============================================================================

st.set_page_config(
    page_title="PDF Data Extractor",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded",
)


# =============================================================================
# CUSTOM CSS
# =============================================================================

st.markdown("""
<style>
    .main .block-container { padding-top: 2rem; max-width: 1200px; }

    .app-header {
        background: linear-gradient(135deg, #1e3a5f 0%, #2f5496 100%);
        color: white; padding: 1.5rem 2rem; border-radius: 12px;
        margin-bottom: 1.5rem; box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }
    .app-header h1 { margin: 0; font-size: 1.8rem; font-weight: 700; }
    .app-header p  { margin: 0.3rem 0 0 0; opacity: 0.85; font-size: 0.95rem; }

    .stat-card {
        background: white; border: 1px solid #e0e7ef; border-radius: 10px;
        padding: 1rem 1.2rem; text-align: center;
        box-shadow: 0 2px 8px rgba(0,0,0,0.04);
    }
    .stat-card .stat-value { font-size: 1.6rem; font-weight: 700; color: #2f5496; }
    .stat-card .stat-label {
        font-size: 0.8rem; color: #6b7280;
        text-transform: uppercase; letter-spacing: 0.5px;
    }

    .file-item {
        background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px;
        padding: 0.6rem 1rem; margin-bottom: 0.4rem;
        display: flex; align-items: center; gap: 0.5rem;
    }

    .success-banner {
        background: linear-gradient(135deg, #059669 0%, #10b981 100%);
        color: white; padding: 1rem 1.5rem; border-radius: 10px; margin: 1rem 0;
    }

    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}

    [data-testid="stFileUploader"] {
        border: 2px dashed #2f5496; border-radius: 12px; padding: 1rem;
    }
</style>
""", unsafe_allow_html=True)


# =============================================================================
# REGEX PATTERNS
# =============================================================================

PO_HEADER = re.compile(
    r"^(4500\d{6})\s+(FO|NB)\s+(\d+)\s+(.+?)\s+(BC\d)\s+(\d{2}/\d{2}/\d{4})"
)
LINE_ITEM = re.compile(r"^(0[0-9]{4})\s+(.+)")
ACCOUNT_LINE = re.compile(
    r"^(?:L\s+)?(?:B\s+)?(\d)\s+([A-Z]{2,4}\d?)\s*(\d{4})?\s+\d+\s+(?:PU|EA|BU)\s+"
    r"([\d,]+\.\d{2})\s+USD"
)
INVOICED = re.compile(
    r"Still to be invoiced\s+(\d+|[OQ]+)\s+(?:PU|EA|BU)\s+"
    r"([\d,]+\.\d{2})\s+USD\s+([\d.]+)\s*[%&$]*"
)
INVOICED_ZERO = re.compile(
    r"Still to be invoiced\s+[OQ]+\s+(?:PU|EA|BU)\s+0\.00\s+USD\s+0\.00\s*[%&$]*"
)


# =============================================================================
# PDF SPLITTING (for PDFs > chunk_size pages)
# =============================================================================

def split_pdf_bytes(pdf_bytes, chunk_size=10):
    """
    If a PDF has more than chunk_size pages, split it into multiple
    byte buffers of chunk_size pages each.
    Returns list of (part_label, pdf_bytes, page_count) tuples.
    """
    reader = PdfReader(io.BytesIO(pdf_bytes))
    num_pages = len(reader.pages)

    if num_pages <= chunk_size:
        return [("full", pdf_bytes, num_pages)]

    parts = []
    part_count = ceil(num_pages / chunk_size)

    for part_idx in range(part_count):
        start = part_idx * chunk_size
        end = min(start + chunk_size, num_pages)

        writer = PdfWriter()
        for p in range(start, end):
            writer.add_page(reader.pages[p])

        buf = io.BytesIO()
        writer.write(buf)
        buf.seek(0)

        label = f"part_{part_idx + 1}_pages_{start + 1}-{end}"
        parts.append((label, buf.getvalue(), end - start))

    return parts


# =============================================================================
# PDF -> Images -> OCR
# =============================================================================

def pdf_bytes_to_text(pdf_bytes, dpi=300):
    """Write PDF bytes to temp file, convert to images, OCR each page."""
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name

    try:
        images = convert_from_path(tmp_path, dpi=dpi)
        page_texts = []
        page_images = []
        for img in images:
            text = pytesseract.image_to_string(img, config="--psm 6")
            page_texts.append(text)
            thumb = img.copy()
            thumb.thumbnail((300, 400))
            page_images.append(thumb)
        return page_texts, page_images
    finally:
        os.unlink(tmp_path)


# =============================================================================
# PARSE OCR TEXT
# =============================================================================

def parse_amount(text):
    return float(text.replace(",", ""))


def parse_pages(page_texts, source_filename="", page_offset=0):
    """Parse OCR text into structured records.
    page_offset is added to page numbers when processing split chunks
    so that page numbers reflect the original PDF.
    """
    records = []
    current = {}
    pending = None

    for page_num, page_text in enumerate(page_texts, 1):
        actual_page = page_num + page_offset
        for raw_line in page_text.splitlines():
            line = raw_line.strip()
            if not line or line.startswith("ee en ed"):
                continue

            m = PO_HEADER.match(line)
            if m:
                current = {
                    "po": m.group(1), "type": m.group(2), "vid": m.group(3),
                    "vname": m.group(4).strip(), "bc": m.group(5), "date": m.group(6),
                }
                continue

            m = LINE_ITEM.match(line)
            if m and current:
                current["line_num"] = m.group(1)
                current["desc"] = m.group(2).strip()
                continue

            m = ACCOUNT_LINE.match(line)
            if m and current:
                acct = f"{m.group(2)} {m.group(3) or ''}".strip()
                pending = {
                    "Source File": source_filename,
                    "Page": actual_page,
                    "PO Number": current.get("po", ""),
                    "PO Type": current.get("type", ""),
                    "Vendor ID": current.get("vid", ""),
                    "Vendor Name": current.get("vname", ""),
                    "Buyer Code": current.get("bc", ""),
                    "PO Date": current.get("date", ""),
                    "Line Item": current.get("line_num", ""),
                    "Description": current.get("desc", ""),
                    "Account Code": acct,
                    "PO Line Amount (USD)": parse_amount(m.group(4)),
                    "Still to be Invoiced (USD)": None,
                    "Invoiced %": None,
                }
                continue

            m = INVOICED.match(line)
            if m and pending:
                pending["Still to be Invoiced (USD)"] = parse_amount(m.group(2))
                pending["Invoiced %"] = float(m.group(3))
                records.append(pending)
                pending = None
                continue

            m = INVOICED_ZERO.match(line)
            if m and pending:
                pending["Still to be Invoiced (USD)"] = 0.00
                pending["Invoiced %"] = 0.00
                records.append(pending)
                pending = None
                continue

    if pending:
        pending["Still to be Invoiced (USD)"] = "N/A (PDF truncated)"
        pending["Invoiced %"] = "N/A (PDF truncated)"
        records.append(pending)

    return records


# =============================================================================
# PROCESS ONE UPLOADED FILE (with auto-splitting)
# =============================================================================

def process_one_file(uploaded_file, dpi=300, chunk_size=10):
    """
    Process a single uploaded PDF:
      1. Check page count
      2. If > chunk_size pages, split into chunks
      3. OCR each chunk
      4. Merge all records together with correct page numbers
    Returns (records, total_pages, all_thumbnails, split_info_str)
    """
    filename = uploaded_file.name
    pdf_bytes = uploaded_file.getvalue()

    # Split if needed
    parts = split_pdf_bytes(pdf_bytes, chunk_size=chunk_size)
    was_split = len(parts) > 1

    all_records = []
    all_thumbnails = []
    total_pages = 0
    page_offset = 0

    split_info = ""
    if was_split:
        total_original_pages = len(PdfReader(io.BytesIO(pdf_bytes)).pages)
        split_info = f"Split into {len(parts)} chunks ({total_original_pages} pages)"

    for part_label, part_bytes, part_page_count in parts:
        page_texts, page_imgs = pdf_bytes_to_text(part_bytes, dpi=dpi)
        records = parse_pages(page_texts, source_filename=filename, page_offset=page_offset)

        all_records.extend(records)
        all_thumbnails.extend(page_imgs)
        total_pages += part_page_count
        page_offset += part_page_count

    return all_records, total_pages, all_thumbnails, split_info


# =============================================================================
# EXCEL WRITING
# =============================================================================

HEADERS_ALL = [
    "Source File", "Page", "PO Number", "PO Type", "Vendor ID", "Vendor Name",
    "Buyer Code", "PO Date", "Line Item", "Description",
    "Account Code", "PO Line Amount (USD)",
    "Still to be Invoiced (USD)", "Invoiced %",
]
HEADERS_SINGLE = [h for h in HEADERS_ALL if h != "Source File"]

COL_WIDTHS_ALL = [30, 6, 14, 8, 10, 38, 10, 12, 10, 35, 14, 20, 24, 12]
COL_WIDTHS_SINGLE = [6, 14, 8, 10, 38, 10, 12, 10, 35, 14, 20, 24, 12]

CENTER_COLS = {"Page", "PO Number", "PO Type", "Vendor ID",
               "Line Item", "Buyer Code", "PO Date"}
MONEY_COLS = {"PO Line Amount (USD)", "Still to be Invoiced (USD)"}

HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="2F5496")
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")
BORDER = Border(
    left=Side("thin", color="D9D9D9"), right=Side("thin", color="D9D9D9"),
    top=Side("thin", color="D9D9D9"), bottom=Side("thin", color="D9D9D9"),
)
BOLD_FONT = Font(name="Arial", bold=True, size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="2F5496")
NORMAL_FONT = Font(name="Arial", size=11)


def write_data_sheet(ws, records, headers, col_widths):
    for c, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=hdr)
        cell.font, cell.fill, cell.alignment, cell.border = HDR_FONT, HDR_FILL, HDR_ALIGN, BORDER

    for r, rec in enumerate(records, 2):
        is_alt = r % 2 == 0
        for c, key in enumerate(headers, 1):
            val = rec.get(key, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.font, cell.border = DATA_FONT, BORDER
            if is_alt:
                cell.fill = ALT_FILL
            if key in MONEY_COLS and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif key == "Invoiced %" and isinstance(val, (int, float)):
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")
            elif key in CENTER_COLS:
                cell.alignment = Alignment(horizontal="center")

    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(records) + 1}"


def write_summary_sheet(ws, file_summaries, grand_totals):
    ws["A1"] = "CONSOLIDATED EXTRACTION SUMMARY"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = NORMAL_FONT

    row = 4
    ws.cell(row=row, column=1, value="GRAND TOTALS").font = Font(
        name="Arial", bold=True, size=13, color="C00000"
    )
    row += 1
    for label, val in [
        ("Total PDF Files Processed", grand_totals["files"]),
        ("Total Pages Processed", grand_totals["pages"]),
        ("Total Unique PO Numbers", grand_totals["unique_pos"]),
        ("Total Line Items Extracted", grand_totals["line_items"]),
        ("Total PO Line Amount (USD)", grand_totals["total_amount"]),
        ("Total Still to be Invoiced (USD)", grand_totals["total_invoiced"]),
    ]:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        c = ws.cell(row=row, column=2, value=val)
        c.font = NORMAL_FONT
        if isinstance(val, float):
            c.number_format = "$#,##0.00"
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="PER-FILE BREAKDOWN").font = Font(
        name="Arial", bold=True, size=13, color="2F5496"
    )
    row += 1
    for c, hdr in enumerate(["File Name", "Pages", "PO Count", "Line Items",
                              "PO Line Total (USD)", "Invoiced Total (USD)", "Anomalies"], 1):
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font, cell.fill, cell.alignment, cell.border = HDR_FONT, HDR_FILL, HDR_ALIGN, BORDER
    row += 1

    for fs in file_summaries:
        for c, val in enumerate([
            fs["filename"], fs["pages"], fs["unique_pos"], fs["line_items"],
            fs["total_amount"], fs["total_invoiced"], fs["anomalies"] or "None"
        ], 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font, cell.border = DATA_FONT, BORDER
            if isinstance(val, float):
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
        row += 1

    for i, w in enumerate([40, 8, 10, 12, 22, 22, 40]):
        ws.column_dimensions[get_column_letter(i + 1)].width = w


def create_single_excel(records, num_pages, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Full Extraction"
    write_data_sheet(ws, records, HEADERS_SINGLE, COL_WIDTHS_SINGLE)

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = f"EXTRACTION SUMMARY - {filename}"
    ws2["A1"].font = TITLE_FONT
    total_amt = sum(r["PO Line Amount (USD)"] for r in records if isinstance(r["PO Line Amount (USD)"], (int, float)))
    total_inv = sum(r["Still to be Invoiced (USD)"] for r in records if isinstance(r["Still to be Invoiced (USD)"], (int, float)))
    unique_pos = len(set(r["PO Number"] for r in records))
    truncated = [r for r in records if r.get("Invoiced %") == "N/A (PDF truncated)"]

    summary_rows = [
        ("Pages Processed", num_pages), ("Total PO Numbers", unique_pos),
        ("Total Line Items", len(records)), ("Total PO Line Amount (USD)", total_amt),
        ("Total Still to be Invoiced (USD)", total_inv), ("", ""), ("ANOMALIES", ""),
    ]
    if truncated:
        for t in truncated:
            summary_rows.append((f"PO {t['PO Number']} truncated", "N/A"))
    else:
        summary_rows.append(("None found", ""))

    for i, (label, val) in enumerate(summary_rows, 3):
        ws2.cell(row=i, column=1, value=label).font = BOLD_FONT
        c = ws2.cell(row=i, column=2, value=val)
        c.font = NORMAL_FONT
        if isinstance(val, float):
            c.number_format = "$#,##0.00"
    ws2.column_dimensions["A"].width = 60
    ws2.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def create_consolidated_excel(all_records, file_summaries, grand_totals):
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "All Data"
    write_data_sheet(ws_all, all_records, HEADERS_ALL, COL_WIDTHS_ALL)

    files_seen = []
    for fs in file_summaries:
        fname = fs["filename"]
        tab_name = re.sub(r'[\\/*?\[\]:]', '', os.path.splitext(fname)[0])[:31]
        base_tab = tab_name
        counter = 1
        while tab_name in files_seen:
            suffix = f"_{counter}"
            tab_name = base_tab[:31 - len(suffix)] + suffix
            counter += 1
        files_seen.append(tab_name)
        file_records = [r for r in all_records if r["Source File"] == fname]
        ws_file = wb.create_sheet(tab_name)
        write_data_sheet(ws_file, file_records, HEADERS_SINGLE, COL_WIDTHS_SINGLE)

    ws_summary = wb.create_sheet("Summary")
    write_summary_sheet(ws_summary, file_summaries, grand_totals)
    wb.move_sheet("Summary", offset=-(len(file_summaries)))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# =============================================================================
# UI COMPONENTS
# =============================================================================

def render_header():
    st.markdown("""
    <div class="app-header">
        <h1>📄 PDF Data Extractor</h1>
        <p>Upload financial PDFs → Auto-split large files → OCR → Extract to Excel</p>
    </div>
    """, unsafe_allow_html=True)


def render_sidebar():
    with st.sidebar:
        st.markdown("### ⚙️ Settings")

        dpi = st.select_slider(
            "OCR Resolution (DPI)",
            options=[150, 200, 250, 300, 400, 500],
            value=300,
            help="Higher DPI = better accuracy but slower",
        )

        chunk_size = st.number_input(
            "Max pages per OCR chunk",
            min_value=5, max_value=50, value=10, step=5,
            help="PDFs with more pages than this are auto-split before OCR to save memory",
        )

        show_preview = st.checkbox("Show page image previews", value=False,
                                   help="Display thumbnails of each page")

        st.markdown("---")
        st.markdown("### 📋 How to Use")
        st.markdown("""
        1. **Upload** one or more PDF files
        2. **Click** "Extract Data"
        3. **Download** individual or consolidated Excel
        """)

        st.markdown("---")
        st.markdown("### 🔀 Auto-Splitting")
        st.markdown(f"""
        PDFs with **more than {chunk_size} pages** are automatically
        split into {chunk_size}-page chunks before OCR. This prevents
        memory issues with large files. Page numbers in the
        Excel output reflect the original PDF.
        """)

        st.markdown("---")
        st.markdown("### 👥 Multi-User")
        st.markdown("""
        Multiple users can access this app simultaneously.
        Each session is independent and private.
        """)

    return dpi, chunk_size, show_preview


def render_stats(grand_totals):
    cols = st.columns(5)
    stats = [
        ("📁 Files", grand_totals["files"]),
        ("📄 Pages", grand_totals["pages"]),
        ("🔢 PO Numbers", grand_totals["unique_pos"]),
        ("📝 Line Items", grand_totals["line_items"]),
        ("💰 Total Amount", f"${grand_totals['total_amount']:,.2f}"),
    ]
    for col, (label, value) in zip(cols, stats):
        with col:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-value">{value}</div>
                <div class="stat-label">{label}</div>
            </div>
            """, unsafe_allow_html=True)


# =============================================================================
# MAIN APP
# =============================================================================

def main():
    render_header()
    dpi, chunk_size, show_preview = render_sidebar()

    # ---- Upload ----
    st.markdown("### 📤 Upload PDF Files")
    uploaded_files = st.file_uploader(
        "Drag and drop PDF files here, or click to browse",
        type=["pdf"],
        accept_multiple_files=True,
        help="Upload one or more financial PDF files",
    )

    if not uploaded_files:
        st.info("👆 Upload one or more PDF files to get started.")
        st.markdown("---")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown("#### 📄 Upload")
            st.markdown("Drag & drop one or many PDFs.")
        with col2:
            st.markdown("#### ⚡ Extract")
            st.markdown("Auto-split → image conversion → OCR → parse.")
        with col3:
            st.markdown("#### 📊 Download")
            st.markdown("Individual + consolidated Excel files.")
        return

    # Show file list
    st.markdown(f"**{len(uploaded_files)} file(s) ready:**")
    for f in uploaded_files:
        st.markdown(
            f'<div class="file-item">📄 <strong>{f.name}</strong> '
            f'({f.size / 1024:.1f} KB)</div>',
            unsafe_allow_html=True,
        )
    st.markdown("")

    # ---- Process ----
    if st.button("🚀 Extract Data", type="primary", use_container_width=True):

        all_records = []
        file_summaries = []
        file_excel_bytes = {}
        file_images = {}
        total_pages = 0
        start_time = time.time()

        progress_bar = st.progress(0, text="Starting extraction...")
        status_text = st.empty()

        for file_idx, uploaded_file in enumerate(uploaded_files):
            filename = uploaded_file.name
            progress_pct = file_idx / len(uploaded_files)
            progress_bar.progress(progress_pct, text=f"Processing {filename}...")
            status_text.markdown(f"⏳ **{filename}** — Checking page count & splitting if needed...")

            try:
                records, num_pages, page_imgs, split_info = process_one_file(
                    uploaded_file, dpi=dpi, chunk_size=chunk_size
                )
                total_pages += num_pages
                file_images[filename] = page_imgs

                if split_info:
                    status_text.markdown(f"⏳ **{filename}** — {split_info}, extracting...")
                else:
                    status_text.markdown(f"⏳ **{filename}** — {num_pages} pages, extracting...")

                if not records:
                    file_summaries.append({
                        "filename": filename, "pages": num_pages, "unique_pos": 0,
                        "line_items": 0, "total_amount": 0.0, "total_invoiced": 0.0,
                        "anomalies": "No parseable records found",
                    })
                    continue

                unique_pos = len(set(r["PO Number"] for r in records))
                total_amt = sum(r["PO Line Amount (USD)"] for r in records
                                if isinstance(r["PO Line Amount (USD)"], (int, float)))
                total_inv = sum(r["Still to be Invoiced (USD)"] for r in records
                                if isinstance(r["Still to be Invoiced (USD)"], (int, float)))
                truncated = [r for r in records if r.get("Invoiced %") == "N/A (PDF truncated)"]
                anomaly_parts = []
                if truncated:
                    anomaly_parts.append("; ".join(
                        f"PO {t['PO Number']} line {t['Line Item']} truncated" for t in truncated
                    ))
                if split_info:
                    anomaly_parts.append(split_info)
                anomaly_text = "; ".join(anomaly_parts)

                file_summaries.append({
                    "filename": filename, "pages": num_pages, "unique_pos": unique_pos,
                    "line_items": len(records), "total_amount": total_amt,
                    "total_invoiced": total_inv, "anomalies": anomaly_text,
                })

                file_excel_bytes[filename] = create_single_excel(records, num_pages, filename)
                all_records.extend(records)

            except Exception as e:
                file_summaries.append({
                    "filename": filename, "pages": 0, "unique_pos": 0,
                    "line_items": 0, "total_amount": 0.0, "total_invoiced": 0.0,
                    "anomalies": f"FAILED: {e}",
                })
                st.error(f"Error processing {filename}: {e}")

        progress_bar.progress(1.0, text="Extraction complete!")
        status_text.empty()
        elapsed = time.time() - start_time

        if not all_records:
            st.error("No records could be extracted from any uploaded file.")
            return

        # ---- Grand Totals ----
        grand_totals = {
            "files": len(uploaded_files),
            "pages": total_pages,
            "unique_pos": len(set(r["PO Number"] for r in all_records)),
            "line_items": len(all_records),
            "total_amount": sum(r["PO Line Amount (USD)"] for r in all_records
                                if isinstance(r["PO Line Amount (USD)"], (int, float))),
            "total_invoiced": sum(r["Still to be Invoiced (USD)"] for r in all_records
                                  if isinstance(r["Still to be Invoiced (USD)"], (int, float))),
        }

        consolidated_bytes = create_consolidated_excel(
            all_records, file_summaries, grand_totals
        )

        # ---- Success ----
        st.markdown(f"""
        <div class="success-banner">
            ✅ <strong>Extraction complete!</strong> Processed {grand_totals['files']} file(s),
            {grand_totals['pages']} pages in {elapsed:.1f}s.
        </div>
        """, unsafe_allow_html=True)

        # ---- Stats ----
        st.markdown("### 📊 Extraction Summary")
        render_stats(grand_totals)
        st.markdown("")

        # ---- Downloads ----
        st.markdown("### 💾 Download Results")

        dl_col1, dl_col2 = st.columns([2, 1])
        with dl_col1:
            st.download_button(
                "⬇️ Download CONSOLIDATED Excel (All Files Combined)",
                data=consolidated_bytes,
                file_name="CONSOLIDATED_extraction.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )

        if len(file_excel_bytes) > 1:
            with dl_col2:
                zip_buf = io.BytesIO()
                with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for fname, data in file_excel_bytes.items():
                        zf.writestr(os.path.splitext(fname)[0] + "_extraction.xlsx", data)
                zip_buf.seek(0)
                st.download_button(
                    "📦 Download All as ZIP",
                    data=zip_buf.getvalue(),
                    file_name="all_extractions.zip",
                    mime="application/zip",
                    use_container_width=True,
                )

        if file_excel_bytes:
            st.markdown("#### Individual Files")
            cols = st.columns(min(len(file_excel_bytes), 3))
            for idx, (fname, data) in enumerate(file_excel_bytes.items()):
                with cols[idx % 3]:
                    st.download_button(
                        f"📄 {fname}",
                        data=data,
                        file_name=os.path.splitext(fname)[0] + "_extraction.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

        # ---- Per-File Detail ----
        st.markdown("### 📋 Per-File Breakdown")
        for fs in file_summaries:
            fname = fs["filename"]
            icon = "🔴" if fs.get("anomalies", "").startswith("FAILED") else (
                "🟡" if fs["line_items"] == 0 else "🟢"
            )
            with st.expander(
                f"{icon} **{fname}** — {fs['line_items']} items, "
                f"${fs['total_amount']:,.2f}", expanded=False
            ):
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Pages", fs["pages"])
                c2.metric("POs", fs["unique_pos"])
                c3.metric("Lines", fs["line_items"])
                c4.metric("Amount", f"${fs['total_amount']:,.2f}")

                if fs["anomalies"]:
                    st.warning(f"Anomalies: {fs['anomalies']}")

                if show_preview and fname in file_images:
                    st.markdown("**Page Previews:**")
                    img_cols = st.columns(min(len(file_images[fname]), 5))
                    for i, img in enumerate(file_images[fname]):
                        with img_cols[i % 5]:
                            st.image(img, caption=f"Page {i+1}", use_container_width=True)

        # ---- Data Preview ----
        st.markdown("### 🔍 Data Preview (First 50 Rows)")
        df = pd.DataFrame(all_records[:50])
        st.dataframe(df, use_container_width=True, height=400)

        # ---- Anomalies ----
        failed = [f for f in file_summaries if f.get("anomalies", "").startswith("FAILED")]
        warns = [f for f in file_summaries if f.get("anomalies") and not f["anomalies"].startswith("FAILED")]
        if failed or warns:
            st.markdown("### ⚠️ Anomalies")
            for f in failed:
                st.error(f"**{f['filename']}**: {f['anomalies']}")
            for f in warns:
                st.warning(f"**{f['filename']}**: {f['anomalies']}")


if __name__ == "__main__":
    main()
